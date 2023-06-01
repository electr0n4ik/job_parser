import openpyxl
import os
from src.abc.abc_job_file import JobFile


class EXCELJobFile(JobFile):
    def __init__(self, filename):
        self.filename = filename

    @staticmethod
    def criteria_matches(vacancy, criteria):
        # Проверяем, соответствует ли зарплата критериям
        if 'salary_min' in criteria and vacancy.get('salary_min') < criteria['salary_min']:
            return False

        # Проверяем, соответствует ли должность критериям
        if 'job_name' in criteria and vacancy.get('job_name') != criteria['job_name']:
            return False
        # TODO
        # Подумать над другими проверками критериев
        # Другие проверки критериев...

        return True

    def add_vacancy_exel(self, vacancy_data):
        os.chdir(os.path.abspath(".."))
        folder_path = os.path.abspath("data_vacancies")
        file_path = os.path.join(folder_path, self.filename)

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Записываем заголовки столбцов
        headers = list(vacancy_data.keys())
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        # Записываем данные в строку
        row_data = list(vacancy_data.values())
        sheet.append(row_data)

        workbook.save(file_path)

    def get_vacancies(self, criteria):
        vacancies = []
        file_path = os.path.join(os.path.abspath("data_vacancies"), self.filename)
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # Читаем данные из всех строк, начиная со второй
            for row in sheet.iter_rows(min_row=2, values_only=True):
                vacancy = dict(zip(sheet[1], row))
                if self.criteria_matches(vacancy, criteria):
                    vacancies.append(vacancy)

        return vacancies

    def remove_vacancy(self, vacancy_id):
        file_path = os.path.join(os.path.abspath("data_vacancies"), self.filename)
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # Ищем вакансию по указанному идентификатору и удаляем строку
            for row in sheet.iter_rows(min_row=2, values_only=True):
                vacancy = dict(zip(sheet[1], row))
                if vacancy.get('id') == vacancy_id:
                    sheet.delete_rows(row[0] + 1)  # +1, потому что строки в Excel начинаются с 1

            workbook.save(file_path)
